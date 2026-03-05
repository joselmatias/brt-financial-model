# =========================================================
#  funciones.py – Lógica de cálculo del flujo de caja
#  Troncal BRT – Arquitectura multi-troncal
# =========================================================

import numpy as np
import pandas as pd


# ─────────────────────────────────────────────
#  UTILIDADES GENÉRICAS
# ─────────────────────────────────────────────

def serie_inflacion(base_anual: float, anios: int, inflacion: float) -> list:
    """Devuelve lista Año1..AñoN aplicando inflación acumulada (t=1 sin inflación)."""
    return [base_anual * ((1.0 + inflacion) ** (t - 1)) for t in range(1, anios + 1)]


def cuota_francesa_mensual(monto: float, tasa_anual: float, plazo_anios: int) -> float:
    """Calcula la cuota mensual de un préstamo con amortización francesa."""
    i = tasa_anual / 12.0
    n = int(plazo_anios * 12)
    if i == 0 or n == 0:
        return monto / n if n > 0 else 0.0
    return monto * i / (1 - (1 + i) ** (-n))


def npv(rate: float, cashflows) -> float:
    """Valor Actual Neto dado una tasa y una serie de flujos (t=0, 1, 2, ...)."""
    return sum(cf / ((1.0 + rate) ** t) for t, cf in enumerate(cashflows))


def irr_biseccion(cashflows, low: float = -0.99, high: float = 10.0,
                  tol: float = 1e-8, max_iter: int = 1_000) -> float:
    """
    Calcula la TIR mediante bisección.
    Retorna np.nan si no converge o no hay cambio de signo.
    """
    def f(r):
        return npv(r, cashflows)

    f_low, f_high = f(low), f(high)
    tries = 0
    while f_low * f_high > 0 and tries < 60:
        high *= 2
        f_high = f(high)
        tries += 1

    if f_low * f_high > 0:
        return np.nan

    for _ in range(max_iter):
        mid = (low + high) / 2.0
        f_mid = f(mid)
        if abs(f_mid) < tol:
            return mid
        if f_low * f_mid < 0:
            high, f_high = mid, f_mid
        else:
            low, f_low = mid, f_mid
    return mid


def _cols_anios(anios: int) -> list:
    return [f"Año {i}" for i in range(1, anios + 1)]


def _add_year0(df: pd.DataFrame) -> pd.DataFrame:
    """Inserta columna 'Año 0' con 0.0 si no existe."""
    if "Año 0" not in df.columns:
        df.insert(0, "Año 0", 0.0)
    return df


def calcular_payback(flujos_0aN: np.ndarray) -> str:
    """
    Devuelve el año en que el flujo acumulado se vuelve positivo.
    Retorna 'No recupera' si nunca ocurre en el horizonte.
    """
    acum = np.cumsum(flujos_0aN)
    for i, v in enumerate(acum):
        if v >= 0:
            return f"Año {i}"
    return "No recupera"


# ─────────────────────────────────────────────
#  DEMANDA
# ─────────────────────────────────────────────

def proyectar_demanda(base_demanda: dict, anios: int,
                      tasas_por_anio: list, modo_redondeo: str = "floor") -> pd.DataFrame:
    """
    Proyecta la demanda anual por categoría.

    Parámetros
    ----------
    base_demanda    : dict  {categoría: pasajeros_año1}
    anios           : int   número de años del horizonte
    tasas_por_anio  : list  (anios-1) tasas de crecimiento
    modo_redondeo   : str   'floor' o 'round'

    Retorna DataFrame con índice = categorías + TOTAL, columnas = Año 1..N + TOTAL
    """
    if len(tasas_por_anio) != anios - 1:
        raise ValueError(
            f"tasas_por_anio debe tener {anios - 1} elementos, "
            f"pero tiene {len(tasas_por_anio)}."
        )

    cols = _cols_anios(anios)
    df = pd.DataFrame(0.0, index=list(base_demanda.keys()), columns=cols)

    for k, v in base_demanda.items():
        df.loc[k, "Año 1"] = float(v)

    for t in range(1, anios):
        factor = 1 + float(tasas_por_anio[t - 1])
        prev_col, curr_col = cols[t - 1], cols[t]
        tmp = df[prev_col] * factor
        if modo_redondeo == "floor":
            df[curr_col] = np.floor(tmp)
        else:
            df[curr_col] = np.round(tmp, 0)

    df.loc["TOTAL DEMANDA"] = df.sum(axis=0)
    df["TOTAL"] = df.sum(axis=1)
    return df


def demanda_equivalente_por_regla(demanda_df: pd.DataFrame,
                                   divisores_equivalencia: dict) -> pd.DataFrame:
    """Divide cada categoría por su divisor de equivalencia y totaliza."""
    cats = [c for c in demanda_df.index if c in divisores_equivalencia]
    cols = [c for c in demanda_df.columns if c.startswith("Año")]
    df_eq = demanda_df.loc[cats, cols].astype(float).copy()
    for cat in cats:
        df_eq.loc[cat] = np.round(df_eq.loc[cat] / divisores_equivalencia[cat], 0)
    df_eq.loc["TOTAL DEMANDA EQUIVALENTE"] = df_eq.sum(axis=0)
    df_eq["TOTAL"] = df_eq.sum(axis=1)
    return df_eq


# ─────────────────────────────────────────────
#  INGRESOS
# ─────────────────────────────────────────────

def calcular_ingresos_por_categoria(demanda_df: pd.DataFrame,
                                    tarifas: dict) -> pd.DataFrame:
    """Multiplica demanda × tarifa por categoría y totaliza."""
    cats = [c for c in demanda_df.index if c in tarifas]
    cols = [c for c in demanda_df.columns if c.startswith("Año")]
    ing = pd.DataFrame(0.0, index=cats, columns=cols)
    for cat in cats:
        ing.loc[cat] = demanda_df.loc[cat, cols].astype(float) * float(tarifas[cat])
    ing.loc["TOTAL INGRESOS"] = ing.sum(axis=0)
    ing["TOTAL"] = ing.sum(axis=1)
    return ing


# ─────────────────────────────────────────────
#  COSTOS VARIABLES
# ─────────────────────────────────────────────

def serie_mantenimiento(km_totales: float, costo_km: float, unidades: int,
                         divisor_meses: int, div_pre_7: float,
                         div_post_7: float, anios: int) -> list:
    """
    Calcula la serie anual de costos de mantenimiento.
    Años 1–6 dividen por div_pre_7, años 7–N por div_post_7.
    """
    base_anual = km_totales * costo_km * unidades / divisor_meses
    return [
        base_anual / div_pre_7 if i <= 6 else base_anual / div_post_7
        for i in range(1, anios + 1)
    ]


def serie_combustible(km_totales_troncal: float, km_totales_alim_12y: float,
                       unidades_troncal: int, unidades_alim: int,
                       rend_troncal: float, rend_alim: float,
                       precio_galon: float, anios: int,
                       aplicar_prorrateo: bool = False,
                       divisor_meses: int = 12,
                       div_pre_7: float = 2.0,
                       div_post_7: float = 1.5) -> tuple:
    """
    Calcula series anuales de costo de combustible para troncal y alimentación.
    Retorna (serie_troncal, serie_alim).
    """
    if rend_troncal == 0 or rend_alim == 0:
        raise ValueError("Los rendimientos de combustible no pueden ser 0.")

    km_anual_troncal = km_totales_troncal / divisor_meses
    km_anual_alim    = km_totales_alim_12y / divisor_meses
    gal_troncal_base = (km_anual_troncal * unidades_troncal) / rend_troncal
    gal_alim_base    = (km_anual_alim    * unidades_alim)    / rend_alim

    serie_t, serie_a = [], []
    for i in range(1, anios + 1):
        f = (1.0 / div_pre_7) if (aplicar_prorrateo and i <= 6) \
            else (1.0 / div_post_7) if aplicar_prorrateo else 1.0
        serie_t.append(gal_troncal_base * f * precio_galon)
        serie_a.append(gal_alim_base    * f * precio_galon)
    return serie_t, serie_a


# ─────────────────────────────────────────────
#  NÚCLEO DE CÁLCULO PRINCIPAL
# ─────────────────────────────────────────────

def calcular_modelo(p: dict) -> dict:
    """
    Ejecuta el modelo completo de flujo de caja para una troncal.

    Parámetros
    ----------
    p : dict  – parámetros de la troncal (estructura de TRONCAL_1_DEFAULT)

    Retorna
    -------
    dict con todos los DataFrames y KPIs listos para la UI.
    """
    # ── Validaciones básicas ─────────────────────────────────────
    anios = int(p["anios"])
    cols  = _cols_anios(anios)

    if any(v <= 0 for v in p["tarifas"].values()):
        raise ValueError("Todas las tarifas deben ser mayores que 0.")
    if p["tasa_descuento"] < 0:
        raise ValueError("La tasa de descuento no puede ser negativa.")
    if p["rend_km_gal_troncal"] == 0 or p["rend_km_gal_alim"] == 0:
        raise ValueError("Los rendimientos de combustible no pueden ser 0.")
    if any(abs(t) > 0.5 for t in p["tasas_por_anio"]):
        raise ValueError("Alguna tasa de crecimiento supera ±50%. Revise los valores.")
    if p["unidades_troncal"] == 0 or p["unidades_alim"] == 0:
        raise ValueError("El número de unidades no puede ser 0.")

    # ── DEMANDA ──────────────────────────────────────────────────
    demanda = proyectar_demanda(
        p["base_demanda"], anios, p["tasas_por_anio"], p["modo_redondeo"]
    )
    demanda_equivalente = demanda_equivalente_por_regla(demanda, p["divisores_equivalencia"])

    # ── INGRESOS ─────────────────────────────────────────────────
    ingresos = calcular_ingresos_por_categoria(
        demanda.loc[list(p["base_demanda"].keys())], p["tarifas"]
    )
    ingresos_equivalentes = calcular_ingresos_por_categoria(
        demanda_equivalente, p["tarifas"]
    )

    serie_ingresos_totales = ingresos.loc["TOTAL INGRESOS", cols].astype(float)

    # ── COSTOS VARIABLES: Mantenimiento ─────────────────────────
    serie_mant_troncal = serie_mantenimiento(
        p["km_totales_troncal"], p["costo_km_troncal"], p["unidades_troncal"],
        p["divisor_meses"], p["div_pre_7"], p["div_post_7"], anios
    )
    serie_mant_alim = serie_mantenimiento(
        p["km_totales_alim_12y"], p["costo_km_alim"], p["unidades_alim"],
        p["divisor_meses"], p["div_pre_7"], p["div_post_7"], anios
    )
    df_mant = pd.DataFrame(
        [serie_mant_troncal, serie_mant_alim],
        index=["Mantenimiento (Troncal 18 m)", "Mantenimiento (Alimentación 12 m)"],
        columns=cols
    ).astype(float)
    df_mant.loc["SUBTOTAL MANTENIMIENTO"] = df_mant.sum(axis=0)

    # ── COSTOS VARIABLES: Combustible ────────────────────────────
    comb_troncal, comb_alim = serie_combustible(
        p["km_totales_troncal"], p["km_totales_alim_12y"],
        p["unidades_troncal"], p["unidades_alim"],
        p["rend_km_gal_troncal"], p["rend_km_gal_alim"],
        p["precio_galon"], anios,
        aplicar_prorrateo=p["combustible_aplica_prorrateo"],
        divisor_meses=p["divisor_meses"],
        div_pre_7=p["div_pre_7"], div_post_7=p["div_post_7"]
    )
    df_comb = pd.DataFrame(
        [comb_troncal, comb_alim],
        index=["Combustible (Troncal 18 m)", "Combustible (Alimentación 12 m)"],
        columns=cols
    ).astype(float)
    df_comb.loc["SUBTOTAL COMBUSTIBLE"] = df_comb.sum(axis=0)

    # ── COSTOS VARIABLES: Neumáticos ─────────────────────────────
    costo_neum_base = (
        p["costo_llanta"] *
        (p["unidades_troncal"] * p["llantas_por_bus_troncal"] +
         p["unidades_alim"]    * p["llantas_por_bus_alim"]) *
        p["renovaciones_llantas_por_anio"]
    )
    df_neum = pd.DataFrame(
        [serie_inflacion(costo_neum_base, anios, p["inflacion_anual"])],
        index=["Neumáticos"], columns=cols
    ).astype(float)

    # ── COSTOS VARIABLES: Consolidado operativos ─────────────────
    costos_variables_op = pd.concat([df_mant, df_comb, df_neum], axis=0)
    costos_variables_op.loc["SUBTOTAL COSTOS VARIABLES (operativos)"] = (
        costos_variables_op.loc[
            ["SUBTOTAL MANTENIMIENTO", "SUBTOTAL COMBUSTIBLE", "Neumáticos"]
        ].sum(axis=0)
    )
    costos_variables_op["TOTAL"] = costos_variables_op.sum(axis=1)
    costos_variables_op = costos_variables_op.round(2)

    # ── OTROS COSTOS: ITOR ───────────────────────────────────────
    serie_oper_recaudo = (serie_ingresos_totales * p["itor_porcentaje_oper_recaudo"]).tolist()
    df_itor = pd.DataFrame(
        [serie_oper_recaudo,
         [p["itor_transporte_valores_anual"]] * anios,
         [p["itor_fideicomiso_admin_anual"]]  * anios],
        index=["Costo de Operación y Recaudo (9.95% ingresos)",
               "Costo de Transporte de Valores",
               "Costo de Fideicomiso Administración"],
        columns=cols
    ).astype(float)
    df_itor.loc["TOTAL OTROS COSTOS (ITOR)"] = df_itor.sum(axis=0)
    df_itor["TOTAL"] = df_itor.sum(axis=1)
    df_itor = df_itor.round(2)

    # ── FEE METROVÍA ─────────────────────────────────────────────
    serie_demanda_total = demanda.loc["TOTAL DEMANDA", cols].astype(float)
    df_fee = pd.DataFrame(
        [(serie_demanda_total * p["fee_metrovia_por_pasajero"]).tolist()],
        index=["Fee Metrovía (0.02 × demanda)"],
        columns=cols
    ).astype(float)
    df_fee.loc["TOTAL FEE METROVÍA"] = df_fee.sum(axis=0)
    df_fee["TOTAL"] = df_fee.sum(axis=1)
    df_fee = df_fee.round(2)

    # ── TOTAL COSTOS VARIABLES ───────────────────────────────────
    serie_cv_total = (
        costos_variables_op.loc["SUBTOTAL COSTOS VARIABLES (operativos)", cols].astype(float)
        + df_itor.loc["TOTAL OTROS COSTOS (ITOR)", cols].astype(float)
        + df_fee.loc["TOTAL FEE METROVÍA", cols].astype(float)
    )
    df_total_cv = pd.DataFrame(
        [serie_cv_total.tolist()], index=["TOTAL COSTOS VARIABLES"], columns=cols
    ).astype(float)
    df_total_cv["TOTAL"] = df_total_cv.sum(axis=1)
    df_total_cv = df_total_cv.round(2)

    # ── COSTOS FIJOS: Financiamiento ─────────────────────────────
    cuota_mensual_troncal = cuota_francesa_mensual(
        p["precio_bus_troncal"] * p["porcentaje_financiado"],
        p["tasa_interes_anual"], p["plazo_anios_financ"]
    )
    cuota_mensual_alim = cuota_francesa_mensual(
        p["precio_bus_alimentador"] * p["porcentaje_financiado"],
        p["tasa_interes_anual"], p["plazo_anios_financ"]
    )
    cuota_total_mensual = (
        cuota_mensual_troncal * p["unidades_troncal"] +
        cuota_mensual_alim    * p["unidades_alim"]
    )
    serie_financ = [
        cuota_total_mensual * 12 if i <= p["plazo_anios_financ"] else 0.0
        for i in range(1, anios + 1)
    ]
    df_financ = pd.DataFrame(
        [serie_financ], index=["Costos de financiamiento"], columns=cols
    ).astype(float)

    # ── COSTOS FIJOS: Sueldos ────────────────────────────────────
    sueldo_base_troncal = (
        p["unidades_troncal"] * p["choferes_por_bus_troncal"] *
        p["salario_mensual"] * 12
    )
    sueldo_base_alim = (
        p["unidades_alim"] * p["choferes_por_bus_alim"] *
        p["salario_mensual"] * 12
    )
    df_sueldos = pd.DataFrame(
        [serie_inflacion(sueldo_base_troncal, anios, p["inflacion_anual"]),
         serie_inflacion(sueldo_base_alim,    anios, p["inflacion_anual"])],
        index=["Sueldos (Troncal 18 m)", "Sueldos (Alimentación 12 m)"],
        columns=cols
    ).astype(float)
    df_sueldos.loc["SUBTOTAL SUELDOS"] = df_sueldos.sum(axis=0)

    # ── COSTOS FIJOS: Gastos Administrativos ─────────────────────
    df_adm_items = pd.DataFrame(p["gastos_adm_items"])
    total_adm_anual_base = float((df_adm_items["cantidad"] * df_adm_items["precio"]).sum()) * 12.0
    df_gastos_adm = pd.DataFrame(
        [serie_inflacion(total_adm_anual_base, anios, p["inflacion_anual"])],
        index=["Gastos Administrativos"], columns=cols
    ).astype(float)

    # ── COSTOS FIJOS: Otros rubros ───────────────────────────────
    seguro_total_anual        = p["seguro_fiel_cumpl"] + p["seguro_todo_riesgo_unidades"]
    serv_basicos_anual        = p["serv_basicos_mensual"] * 12
    total_buses               = int(p["unidades_troncal"] + p["unidades_alim"])
    matricula_total           = total_buses * p["matricula_precio"]
    seg_unid_total_anual      = total_buses * p["seg_unid_precio_mensual"] * 12
    matric_impuestos_anual    = matricula_total + p["iva_compras"] + seg_unid_total_anual

    df_seguro      = pd.DataFrame([[seguro_total_anual]     * anios], index=["Seguro"],                     columns=cols).astype(float)
    df_serv        = pd.DataFrame([[serv_basicos_anual]     * anios], index=["Servicios básicos"],          columns=cols).astype(float)
    df_matric      = pd.DataFrame([[matric_impuestos_anual] * anios], index=["Matrícula e impuestos"],      columns=cols).astype(float)
    df_otros_adm   = pd.DataFrame(
        [serie_inflacion(p["otros_adm_anual"], anios, p["inflacion_anual"])],
        index=["Otros costos administrativos"], columns=cols
    ).astype(float)

    # ── COSTOS FIJOS: Consolidado ─────────────────────────────────
    costos_fijos = pd.concat(
        [df_financ, df_sueldos, df_gastos_adm, df_seguro, df_serv, df_matric, df_otros_adm],
        axis=0
    )
    costos_fijos.loc["TOTAL COSTOS FIJOS"] = costos_fijos.loc[
        ["Costos de financiamiento", "SUBTOTAL SUELDOS", "Gastos Administrativos",
         "Seguro", "Servicios básicos", "Matrícula e impuestos",
         "Otros costos administrativos"]
    ].sum(axis=0)
    costos_fijos["TOTAL"] = costos_fijos.sum(axis=1)
    costos_fijos = costos_fijos.round(2)

    # ── COSTOS TOTALES ────────────────────────────────────────────
    serie_costos_totales = (
        df_total_cv.loc["TOTAL COSTOS VARIABLES", cols].astype(float) +
        costos_fijos.loc["TOTAL COSTOS FIJOS", cols].astype(float)
    )
    df_costos_totales = pd.DataFrame(
        [serie_costos_totales.tolist()], index=["COSTOS TOTALES"], columns=cols
    ).astype(float).round(2)
    df_costos_totales["TOTAL"] = df_costos_totales.sum(axis=1)

    # ── UTILIDAD E IMPUESTO ───────────────────────────────────────
    serie_ub = (
        serie_ingresos_totales -
        df_costos_totales.loc["COSTOS TOTALES", cols].astype(float)
    )
    df_utilidad_bruta = pd.DataFrame(
        [serie_ub.tolist()], index=["UTILIDAD BRUTA"], columns=cols
    ).astype(float).round(2)
    df_utilidad_bruta["TOTAL"] = df_utilidad_bruta.sum(axis=1)

    serie_ir = np.where(serie_ub > 0, serie_ub * p["tasa_impuesto_renta"], 0.0)
    df_imp_renta = pd.DataFrame(
        [serie_ir.tolist()],
        index=["Impuesto a la renta (25% util. > 0)"],
        columns=cols
    ).astype(float).round(2)
    df_imp_renta["TOTAL"] = df_imp_renta.sum(axis=1)

    # ── FLUJO DE CAJA ─────────────────────────────────────────────
    # T1: equity solo sobre alimentadores; T2: sobre troncal + alimentadores
    if p.get("equity_incluye_troncal", False):
        base_equity = (
            p["precio_bus_troncal"]    * p["unidades_troncal"] +
            p["precio_bus_alimentador"] * p["unidades_alim"]
        )
    else:
        base_equity = p["precio_bus_alimentador"] * p["unidades_alim"]
    aporte_equity = base_equity * p["porcentaje_equity"]
    serie_flujo = (
        serie_ub - df_imp_renta.loc["Impuesto a la renta (25% util. > 0)", cols].astype(float)
    ).tolist()
    df_flujo = pd.DataFrame(
        [[-float(aporte_equity)] + serie_flujo],
        index=["FLUJO DE CAJA"],
        columns=["Año 0"] + cols
    ).astype(float).round(2)
    df_flujo["TOTAL"] = df_flujo.loc[
        "FLUJO DE CAJA",
        [c for c in df_flujo.columns if c.startswith("Año")]
    ].sum()

    # ── FLUJO ACUMULADO ───────────────────────────────────────────
    cols_0aN = [c for c in df_flujo.columns if c.startswith("Año")]
    flujos_0aN = df_flujo.loc["FLUJO DE CAJA", cols_0aN].astype(float).values
    df_flujo_acum = pd.DataFrame(
        [np.cumsum(flujos_0aN).tolist()],
        index=["FLUJO DE CAJA ACUMULADO"],
        columns=cols_0aN
    ).astype(float).round(2)
    df_flujo_acum["TOTAL"] = flujos_0aN.sum()

    # ── AÑADIR AÑO 0 A TABLAS SIN ÉL ────────────────────────────
    for df in [demanda, demanda_equivalente, ingresos, ingresos_equivalentes,
               costos_variables_op, df_itor, df_fee, df_total_cv,
               costos_fijos, df_costos_totales, df_utilidad_bruta, df_imp_renta]:
        _add_year0(df)

    # ── KPIs ──────────────────────────────────────────────────────
    van   = npv(p["tasa_descuento"], flujos_0aN)
    tir   = irr_biseccion(flujos_0aN)
    flujo_ultimo = flujos_0aN[-1]
    payback      = calcular_payback(flujos_0aN)

    return {
        # DataFrames
        "demanda":              demanda,
        "demanda_equivalente":  demanda_equivalente,
        "ingresos":             ingresos,
        "ingresos_equivalentes": ingresos_equivalentes,
        "costos_variables_op":  costos_variables_op,
        "df_itor":              df_itor,
        "df_fee":               df_fee,
        "df_total_cv":          df_total_cv,
        "costos_fijos":         costos_fijos,
        "df_costos_totales":    df_costos_totales,
        "df_utilidad_bruta":    df_utilidad_bruta,
        "df_imp_renta":         df_imp_renta,
        "df_flujo":             df_flujo,
        "df_flujo_acum":        df_flujo_acum,
        # Series para gráficos (sin columna TOTAL, con Año 0)
        "cols_0aN":             cols_0aN,
        "cols_anios":           cols,
        "flujos_0aN":           flujos_0aN,
        "serie_ingresos":       ingresos.loc["TOTAL INGRESOS", cols].astype(float).values,
        "serie_costos":         df_costos_totales.loc["COSTOS TOTALES", cols].astype(float).values,
        "serie_cv":             df_total_cv.loc["TOTAL COSTOS VARIABLES", cols].astype(float).values,
        "serie_cf":             costos_fijos.loc["TOTAL COSTOS FIJOS", cols].astype(float).values,
        # KPIs
        "van":          van,
        "tir":          tir,
        "flujo_ultimo": flujo_ultimo,
        "payback":      payback,
    }


# ─────────────────────────────────────────────
#  CONSOLIDADO (T1 + T2 + T3 + T4)
# ─────────────────────────────────────────────

def calcular_consolidado(troncales_params: dict) -> dict:
    """
    Corre calcular_modelo() para cada troncal y consolida los resultados.
    Los parámetros comunes (tarifas, tasas, fee, etc.) se toman de la primera troncal.
    """
    resultados = {nombre: calcular_modelo(p) for nombre, p in troncales_params.items()}

    p0   = list(troncales_params.values())[0]
    anios = int(p0["anios"])
    cols  = _cols_anios(anios)
    cols_0aN = ["Año 0"] + cols
    tarifas  = p0["tarifas"]
    cats     = list(p0["base_demanda"].keys())

    # ── DEMANDA consolidada ───────────────────────────────────────
    demanda_cons = pd.DataFrame(0.0, index=cats, columns=cols)
    for res in resultados.values():
        demanda_cons += res["demanda"].loc[cats, cols].astype(float)
    demanda_cons.loc["TOTAL DEMANDA"] = demanda_cons.sum(axis=0)
    demanda_cons["TOTAL"] = demanda_cons.sum(axis=1)
    serie_dem_cons = demanda_cons.loc["TOTAL DEMANDA", cols].astype(float)
    _add_year0(demanda_cons)

    # ── DEMANDA EQUIVALENTE consolidada ──────────────────────────
    demanda_equiv_cons = pd.DataFrame(0.0, index=cats, columns=cols)
    for res in resultados.values():
        de = res["demanda_equivalente"]
        avail = [c for c in cats if c in de.index]
        demanda_equiv_cons.loc[avail] += de.loc[avail, cols].astype(float)
    demanda_equiv_cons.loc["TOTAL DEMANDA EQUIVALENTE"] = demanda_equiv_cons.sum(axis=0)
    demanda_equiv_cons["TOTAL"] = demanda_equiv_cons.sum(axis=1)
    _add_year0(demanda_equiv_cons)

    # ── INGRESOS consolidados ─────────────────────────────────────
    ingresos_cons       = calcular_ingresos_por_categoria(demanda_cons.loc[cats], tarifas)
    ingresos_equiv_cons = calcular_ingresos_por_categoria(demanda_equiv_cons.loc[cats], tarifas)
    _add_year0(ingresos_cons)
    _add_year0(ingresos_equiv_cons)
    serie_ing_cons = ingresos_cons.loc["TOTAL INGRESOS", cols].astype(float)

    # ── FEE METROVÍA consolidada ──────────────────────────────────
    df_fee_c = pd.DataFrame(
        [(serie_dem_cons * p0["fee_metrovia_por_pasajero"]).tolist()],
        index=["Fee Metrovía (0.02 × demanda)"], columns=cols
    ).astype(float).round(2)
    df_fee_c.loc["TOTAL FEE METROVÍA"] = df_fee_c.sum(axis=0)
    df_fee_c["TOTAL"] = df_fee_c.sum(axis=1)
    _add_year0(df_fee_c)

    # ── ITOR consolidado ──────────────────────────────────────────
    itor_op  = (serie_ing_cons * p0["itor_porcentaje_oper_recaudo"]).tolist()
    itor_tv  = sum(res["df_itor"].loc["Costo de Transporte de Valores",       cols].astype(float) for res in resultados.values()).tolist()
    itor_fid = sum(res["df_itor"].loc["Costo de Fideicomiso Administración",  cols].astype(float) for res in resultados.values()).tolist()
    df_itor_c = pd.DataFrame(
        [itor_op, itor_tv, itor_fid],
        index=["Costo de Operación y Recaudo (9.95% ingresos)",
               "Costo de Transporte de Valores",
               "Costo de Fideicomiso Administración"],
        columns=cols
    ).astype(float).round(2)
    df_itor_c.loc["TOTAL OTROS COSTOS (ITOR)"] = df_itor_c.sum(axis=0)
    df_itor_c["TOTAL"] = df_itor_c.sum(axis=1)
    _add_year0(df_itor_c)

    # ── COSTOS VARIABLES OPERATIVOS consolidados ──────────────────
    cv_rows = [
        "Mantenimiento (Troncal 18 m)", "Mantenimiento (Alimentación 12 m)", "SUBTOTAL MANTENIMIENTO",
        "Combustible (Troncal 18 m)",   "Combustible (Alimentación 12 m)",   "SUBTOTAL COMBUSTIBLE",
        "Neumáticos",
    ]
    costos_variables_op_c = pd.DataFrame(0.0, index=cv_rows, columns=cols)
    for res in resultados.values():
        for row in cv_rows:
            if row in res["costos_variables_op"].index:
                costos_variables_op_c.loc[row] += res["costos_variables_op"].loc[row, cols].astype(float)
    serie_cv_oper = sum(
        res["costos_variables_op"].loc["SUBTOTAL COSTOS VARIABLES (operativos)", cols].astype(float)
        for res in resultados.values()
    )
    costos_variables_op_c.loc["SUBTOTAL COSTOS VARIABLES (operativos)"] = serie_cv_oper
    costos_variables_op_c["TOTAL"] = costos_variables_op_c.sum(axis=1)
    costos_variables_op_c = costos_variables_op_c.round(2)
    _add_year0(costos_variables_op_c)

    # ── TOTAL COSTOS VARIABLES ────────────────────────────────────
    serie_cv_total = (
        serie_cv_oper
        + df_itor_c.loc["TOTAL OTROS COSTOS (ITOR)", cols].astype(float)
        + df_fee_c.loc["TOTAL FEE METROVÍA", cols].astype(float)
    )
    df_total_cv_c = pd.DataFrame(
        [serie_cv_total.tolist()], index=["TOTAL COSTOS VARIABLES"], columns=cols
    ).astype(float).round(2)
    df_total_cv_c["TOTAL"] = df_total_cv_c.sum(axis=1)
    _add_year0(df_total_cv_c)

    # ── COSTOS FIJOS consolidados ─────────────────────────────────
    serie_cf = sum(
        res["costos_fijos"].loc["TOTAL COSTOS FIJOS", cols].astype(float)
        for res in resultados.values()
    )
    costos_fijos_c = pd.DataFrame(
        [serie_cf.tolist()], index=["TOTAL COSTOS FIJOS"], columns=cols
    ).astype(float).round(2)
    costos_fijos_c["TOTAL"] = costos_fijos_c.sum(axis=1)
    _add_year0(costos_fijos_c)

    # ── COSTOS TOTALES ─────────────────────────────────────────────
    serie_ct = (
        df_total_cv_c.loc["TOTAL COSTOS VARIABLES", cols].astype(float)
        + costos_fijos_c.loc["TOTAL COSTOS FIJOS", cols].astype(float)
    )
    df_costos_totales_c = pd.DataFrame(
        [serie_ct.tolist()], index=["COSTOS TOTALES"], columns=cols
    ).astype(float).round(2)
    df_costos_totales_c["TOTAL"] = df_costos_totales_c.sum(axis=1)
    _add_year0(df_costos_totales_c)

    # ── UTILIDAD E IMPUESTO ───────────────────────────────────────
    serie_ub = serie_ing_cons - serie_ct
    df_utilidad_bruta_c = pd.DataFrame(
        [serie_ub.tolist()], index=["UTILIDAD BRUTA"], columns=cols
    ).astype(float).round(2)
    df_utilidad_bruta_c["TOTAL"] = df_utilidad_bruta_c.sum(axis=1)
    _add_year0(df_utilidad_bruta_c)

    serie_ir = np.where(serie_ub > 0, serie_ub * p0["tasa_impuesto_renta"], 0.0)
    df_imp_renta_c = pd.DataFrame(
        [serie_ir.tolist()],
        index=["Impuesto a la renta (25% util. > 0)"], columns=cols
    ).astype(float).round(2)
    df_imp_renta_c["TOTAL"] = df_imp_renta_c.sum(axis=1)
    _add_year0(df_imp_renta_c)

    # ── FLUJO DE CAJA ─────────────────────────────────────────────
    aporte_equity_0 = sum(
        float(res["df_flujo"].loc["FLUJO DE CAJA", "Año 0"])
        for res in resultados.values()
    )  # ya es negativo (suma de los 4)
    serie_flujo = (serie_ub - pd.Series(serie_ir, index=cols)).tolist()
    df_flujo_c = pd.DataFrame(
        [[aporte_equity_0] + serie_flujo],
        index=["FLUJO DE CAJA"],
        columns=["Año 0"] + cols
    ).astype(float).round(2)
    df_flujo_c["TOTAL"] = df_flujo_c.loc["FLUJO DE CAJA", cols_0aN].sum()

    # ── FLUJO ACUMULADO ───────────────────────────────────────────
    flujos_0aN = df_flujo_c.loc["FLUJO DE CAJA", cols_0aN].astype(float).values
    df_flujo_acum_c = pd.DataFrame(
        [np.cumsum(flujos_0aN).tolist()],
        index=["FLUJO DE CAJA ACUMULADO"],
        columns=cols_0aN
    ).astype(float).round(2)
    df_flujo_acum_c["TOTAL"] = flujos_0aN.sum()

    # ── KPIs ──────────────────────────────────────────────────────
    van     = npv(p0["tasa_descuento"], flujos_0aN)
    tir     = irr_biseccion(flujos_0aN)
    payback = calcular_payback(flujos_0aN)

    return {
        "demanda":               demanda_cons,
        "demanda_equivalente":   demanda_equiv_cons,
        "ingresos":              ingresos_cons,
        "ingresos_equivalentes": ingresos_equiv_cons,
        "costos_variables_op":   costos_variables_op_c,
        "df_itor":               df_itor_c,
        "df_fee":                df_fee_c,
        "df_total_cv":           df_total_cv_c,
        "costos_fijos":          costos_fijos_c,
        "df_costos_totales":     df_costos_totales_c,
        "df_utilidad_bruta":     df_utilidad_bruta_c,
        "df_imp_renta":          df_imp_renta_c,
        "df_flujo":              df_flujo_c,
        "df_flujo_acum":         df_flujo_acum_c,
        # Series para gráficos
        "cols_0aN":       cols_0aN,
        "cols_anios":     cols,
        "flujos_0aN":     flujos_0aN,
        "serie_ingresos": ingresos_cons.loc["TOTAL INGRESOS", cols].astype(float).values,
        "serie_costos":   df_costos_totales_c.loc["COSTOS TOTALES", cols].astype(float).values,
        "serie_cv":       df_total_cv_c.loc["TOTAL COSTOS VARIABLES", cols].astype(float).values,
        "serie_cf":       costos_fijos_c.loc["TOTAL COSTOS FIJOS", cols].astype(float).values,
        # KPIs
        "van":          van,
        "tir":          tir,
        "flujo_ultimo": flujos_0aN[-1],
        "payback":      payback,
    }


# ─────────────────────────────────────────────
#  ANÁLISIS DE SENSIBILIDAD
# ─────────────────────────────────────────────

def tarifa_general_van_cero(precio_galon: float, troncales_params: dict,
                              tol_usd: float = 1.0) -> float:
    """
    Bisección: encuentra la tarifa GENERAL (aplicada a todas las troncales)
    que hace VAN consolidado = 0, dado un precio_galon fijo.
    Devuelve np.nan si no existe cruce de cero en el rango [0.01, 10.0].
    """
    import copy

    def van_para(tarifa: float) -> float:
        params = copy.deepcopy(troncales_params)
        for p in params.values():
            p["precio_galon"] = precio_galon
            p["tarifas"]["GENERAL"] = tarifa
        return calcular_consolidado(params)["van"]

    low, high = 0.01, 10.0
    f_low, f_high = van_para(low), van_para(high)

    if f_low * f_high > 0:
        return np.nan  # No hay cruce de cero en el rango

    for _ in range(50):
        mid = (low + high) / 2.0
        f_mid = van_para(mid)
        if abs(f_mid) < tol_usd:
            return mid
        if f_low * f_mid < 0:
            high = mid
            f_high = f_mid
        else:
            low = mid
            f_low = f_mid
    return (low + high) / 2.0


def tarifa_general_van_cero_troncal(precio_galon: float, params_troncal: dict,
                                      tol_usd: float = 1.0) -> float:
    """
    Bisección: encuentra la tarifa GENERAL que hace VAN = 0 para UNA troncal,
    dado un precio_galon fijo. Devuelve np.nan si no hay cruce de cero.
    """
    import copy

    def van_para(tarifa: float) -> float:
        p = copy.deepcopy(params_troncal)
        p["precio_galon"] = precio_galon
        p["tarifas"]["GENERAL"] = tarifa
        return calcular_modelo(p)["van"]

    low, high = 0.01, 10.0
    f_low, f_high = van_para(low), van_para(high)

    if f_low * f_high > 0:
        return np.nan

    for _ in range(50):
        mid = (low + high) / 2.0
        f_mid = van_para(mid)
        if abs(f_mid) < tol_usd:
            return mid
        if f_low * f_mid < 0:
            high = mid
            f_high = f_mid
        else:
            low = mid
            f_low = f_mid
    return (low + high) / 2.0


# ─────────────────────────────────────────────
#  EXPORTACIÓN A EXCEL
# ─────────────────────────────────────────────

def exportar_excel(resultado: dict, nombre_troncal: str = "Troncal 1") -> bytes:
    """
    Genera un archivo Excel multi-hoja con todos los resultados.
    Retorna el contenido como bytes para descarga en Streamlit.
    """
    import io
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    hojas = {
        "Demanda":              resultado["demanda"],
        "Demanda_equivalente":  resultado["demanda_equivalente"],
        "Ingresos":             resultado["ingresos"],
        "Ing_equivalentes":     resultado["ingresos_equivalentes"],
        "Costos_variables":     resultado["costos_variables_op"],
        "ITOR":                 resultado["df_itor"],
        "Fee_Metrovia":         resultado["df_fee"],
        "Total_CV":             resultado["df_total_cv"],
        "Costos_fijos":         resultado["costos_fijos"],
        "Costos_totales":       resultado["df_costos_totales"],
        "Utilidad":             resultado["df_utilidad_bruta"],
        "Impuesto_Renta":       resultado["df_imp_renta"],
        "Flujo_de_Caja":        resultado["df_flujo"],
        "Flujo_Acumulado":      resultado["df_flujo_acum"],
    }
    hojas_enteras = {"Demanda", "Demanda_equivalente"}

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for nombre, df in hojas.items():
            df.to_excel(writer, sheet_name=nombre)
            ws = writer.sheets[nombre]
            fmt = "#,##0" if nombre in hojas_enteras else "#,##0.00"

            # Encabezado
            header_fill = PatternFill("solid", fgColor="1F4E79")
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Formato de números
            for row in ws.iter_rows(min_row=2, min_col=2,
                                    max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = fmt

            # Ancho de columnas
            for j in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(j)].width = 16

        # Hoja KPIs
        van  = resultado["van"]
        tir  = resultado["tir"]
        kpis = pd.DataFrame({
            "Indicador": ["Troncal", "VAN (12%)", "TIR", f"Flujo Año {resultado['df_flujo'].shape[1]-2}", "Payback"],
            "Valor":     [
                nombre_troncal,
                f"${van:,.2f}",
                "N/A" if np.isnan(tir) else f"{tir*100:.2f}%",
                f"${resultado['flujo_ultimo']:,.2f}",
                resultado["payback"],
            ]
        })
        kpis.to_excel(writer, sheet_name="KPIs", index=False)

    buffer.seek(0)
    return buffer.read()
